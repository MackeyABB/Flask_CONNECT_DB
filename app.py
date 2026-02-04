'''
LastEditors: Mackey
LastEditTime: 2024-04-29 15:30:00
Introduction: 
    Flask网页端访问CONNECT DB并显示
    即制作网页版的CONNECT DB显示程序

Revision History:
see: Revision_Log.md
'''

# 版本号
# xx.yy.zz
# xx: 大版本，架构性变化
# yy: 功能性新增
# zz: Bug修复
__Version__ = "4.1.0"

import sys
from flask import Flask, send_file , jsonify , request, redirect
from flask import render_template
from flask.helpers import flash, url_for
import pandas as pd  
import openpyxl
import io  
import db_mgt
import re  
import requests  
import base64 
import json  
import logging
import datetime
from werkzeug.utils import secure_filename
import os
import third_party.PLM_Handle.PLM_Basic_Auth_ByPass_MFA_Get_BOM as plm
import third_party.Excel_Handle.AVL_Excel_Handle as excel_handle
import openpyxl
import tempfile
import os
import threading
import time
import win32com.client as win32
from flask import g

# Global Variables
first_AVL_Output_File = ""  # Excel输出文件路径
AVL_Compare_Output_File = ""  # AVL比较输出文件路径

# debug print, print到控制台
DEBUG_PRINT = True

def debug_print(*args, **kwargs):
    if DEBUG_PRINT:
        print(*args, **kwargs)

# production logging
# from logging.config import dictConfig

# dictConfig({
#     'version': 1,
#     'formatters': {'default': {
#         'format': '[%(asctime)s] %(levelname)s in %(module)s: %(message)s',
#     }},
#     'handlers': {
#         'wsgi': {
#             'class': 'logging.StreamHandler',
#             'formatter': 'default'
#         },
#         'custom_handler': {
#             'class': 'logging.FileHandler',
#             'formatter': 'default',
#             'filename': r'D:\80_MackeyDoc\01_ABB\OneDrive - ABB\00_WorkPlace\01_Design_Work\01_Prg\10_Py\PythonCode\A1_Flask\Flask_CONNECT_DB\myapp.log'
#         }
#     },
#     'root': {
#         'level': 'DEBUG',
#         'handlers': ['wsgi', 'custom_handler']
#     }
# })



app = Flask(__name__)

# 对于IIS生产系统,这段要放在这里,不能放在main里
# create DB instance
# db = db_mgt.Database()
def get_db():
    if 'db' not in g:
        g.db = db_mgt.Database()
    return g.db

# set debug, can work even in production
app.config['ENV'] = 'development'
app.config['DEBUG'] = True
app.config['TESTING'] = True

app.secret_key = 'CONNECT'

# Set the loggin
# handler = logging.FileHandler('app.log')  # errors logged to this file
# handler.setLevel(logging.NOTSET)  # only log errors and above
# app.logger.addHandler(handler)  # attach the handler to the app's logger
# app.logger.info("App start !")

# create DB instance
WC_Path = "https://lp-global-plm.abb.com/Windchill/protocolAuth/servlet/odata/"
#定义一个空的集合用于记录AVL里的元器件清单
AVLPart_ListView=set()
#定义一个空的集合用于在网页端显示内容以供使用者检查
Component_ListView=set()
#定义Excel模板中,有效数据的首行
Excel_Row=7

from flask import render_template_string
try:
    import markdown
except ImportError:
    markdown = None
# ...existing code...

# Markdown预览路由
@app.route('/Revision_Log.md/preview')
def revision_log_preview():
    md_path = 'Revision_Log.md'
    if markdown is None:
        return 'Markdown package not installed. Please run: pip install markdown', 500
    try:
        with open(md_path, 'r', encoding='utf-8') as f:
            md_content = f.read()
        html_content = markdown.markdown(md_content, extensions=['tables'])
        style = '''<style>body{font-family:Arial,Helvetica,sans-serif;padding:20px;}table{border-collapse:collapse;}th,td{border:1px solid #888;padding:4px 8px;}th{background:#eee;}</style>'''
        return render_template_string(f'{style}{html_content}')
    except Exception as e:
        return f'Error loading Revision_Log.md: {e}', 404

'''
初始始页面： 选择数据库
点击确认之后,会跳转到检索页面,并且会将选择的数据库序号传递过去。
'''
@app.route("/", methods=['GET','POST'])
def DBSelect():
    DB_List=db_mgt.DBList   
    
    if request.method == "POST":
        # 获取选择的数据库类型, 为str类型
        DB_Select = request.form.get("DB_Select")
        DB_Index = db_mgt.DBList.index(DB_Select)
        print (DB_Index)
        return redirect(url_for('index',DBType=DB_Index))
    else:
        return render_template('DBSelect.html', DB_List=DB_List,Version = __Version__)

'''
数据库检索和显示页面：初始始显示检索表单,在此之前会尝试打开数据库
submit之后显示检索内容
'''
@app.route("/search/<DBType>", methods=['GET','POST'])
def index(DBType):
    db = get_db()
    # 根据DBType来设置Part Type 列表的内容,DBType为str,对应db_mgt.DBList的index值,从0开始
    if DBType == '0' or DBType == '3': 
        #如果将值直接在render_template里赋值,数据第一次会传递不过去,不知原因。
        Part_Type_List=db_mgt.PartTypeList_CONNECT
    elif DBType == '1':
        Part_Type_List=db_mgt.PartTypeList_Access
    else:
        Part_Type_List=db_mgt.PartTypeList_Access
    
    # 尝试打开DB
    # Open DB
    bIsDBOpen = db.openDB(int(DBType), db_mgt.DBList, app)
    if bIsDBOpen == True:
        flash(db_mgt.DBList[int(DBType)]+" Database opened successfully!")
    else:
        flash(db_mgt.DBList[int(DBType)]+" Database open error")
    
    # 提交表单查询时处理
    if request.method == "POST":
        # 定义全局变量
        global sql_result
        global columnNameList
        global MaxLine
        global sql_result_len
        dbindex = int(DBType)      
        # 获取表单内容
        tableName  = request.form.get("tableName")        
        PartNo_Searchby = request.form.get("PartNo")
        SAPNo_Searchby = request.form.get("SAPNo")
        PartValue_Searchby = request.form.get("PartValue")
        MfcPartNum_Searchby = request.form.get("MfcPartNum")
        MaxLine_str = request.form.get("MaxLine")
        MaxLine = int(MaxLine_str) if MaxLine_str is not None and MaxLine_str.strip() != "" else 100  # default to 100 if not provided
        Description_Searchby = request.form.get("Description")
        TechDescription_Searchby = request.form.get("TechDescription")
        Editor_Searchby = request.form.get("Editor")        
        SAP_Number_List = request.form.get("SAP_Number_List")
        Manufacturer_Searchby = request.form.get("Manufacturer")
            
                  
        # 判断按键
        print("press the button:" + request.form['btn'])
        if request.form['btn'] == 'Search':
            # 条件搜索
            Search_Info = f"PartNo: {PartNo_Searchby}, SAPNo: {SAPNo_Searchby}, SAP_Des:{Description_Searchby},PartValue: {PartValue_Searchby},Manufacturer: {Manufacturer_Searchby}, MfcPartNum: {MfcPartNum_Searchby}, TechDescription: {TechDescription_Searchby}, Editor: {Editor_Searchby}, MaxLine: {MaxLine}, TableName: {tableName}"
            # 执行搜索
            sql_result, columnNameList = db.fetch(
                tableName=tableName, 
                dbindex=dbindex, 
                PartNo_Searchby=PartNo_Searchby, 
                SAPNo_Searchby=SAPNo_Searchby, 
                PartValue_Searchby=PartValue_Searchby, 
                MfcPartNum_Searchby=MfcPartNum_Searchby, 
                Description_Searchby=Description_Searchby, 
                TechDescription_Searchby=TechDescription_Searchby, 
                Editor_Searchby=Editor_Searchby,
                Manufacturer_Searchby=Manufacturer_Searchby
            )
            sql_result_len = len(sql_result)
            # 显示结果
            return render_template(
                'index.html', 
                Part_Type_List=Part_Type_List, 
                MaxLine=MaxLine, 
                sql_result=sql_result, 
                columnNameList=columnNameList, 
                sql_result_len=sql_result_len, 
                Search_Info=Search_Info,
                PartNo=PartNo_Searchby,
                SAPNo=SAPNo_Searchby,
                PartValue=PartValue_Searchby,
                MfcPartNum=MfcPartNum_Searchby,
                tableName=tableName,
                Description=Description_Searchby,
                TechDescription=TechDescription_Searchby,
                Manufacturer=Manufacturer_Searchby,
                Editor=Editor_Searchby,
                Version=__Version__
                )
            # return db_mgt.DBList[0]
        elif request.form['btn'] == 'SaveExcel':
            # 保存Excel
            print("SaveExcel")
            # print(columnNameList)
            if 'columnNameList' in globals():
                temp_dir = tempfile.gettempdir()    # not used, as in the server the temp dir is not the same as in the local
                # file_path = os.path.join(os.path.dirname(os.path.realpath(sys.argv[0])), 'ExportFiles', "SQL_Result.xlsx")
                # import datetime
                # 用唯一文件名（如加时间戳）,避免冲突
                filename = f"SQL_Result_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                file_path = os.path.join(os.path.dirname(os.path.realpath(sys.argv[0])), 'ExportFiles', filename)
                if file_path:
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.append(columnNameList)
                    for row in sql_result:
                        ws.append(row)
                    wb.save(file_path)
                    flash("Excel保存成功！{}".format(file_path))
                    # 打开Excel, 文件会保存在服务器中,客户端是无法直接打开这个文件的,此方法行不通的。
                    # print(file_path)                    
                    # os.system('start excel.exe {}'.format('"' + file_path + '"'))
                # 可以使用send_file来发送文件给客户端
                flash("Excel保存成功！")
                return send_file(file_path, as_attachment=True)
                # return render_template('index.html', Part_Type_List=Part_Type_List, MaxLine=MaxLine, sql_result=sql_result, columnNameList=columnNameList, sql_result_len=sql_result_len)
            else:
                flash("没有数据,无法保存Excel！")
                return render_template('index.html', Part_Type_List=Part_Type_List,Version=__Version__)
        elif request.form['btn'] == 'SAP_Nums_Search':
            # 搜索多个SAP编号
            # 处理多个SAP编号的输入,按空格、逗号、分号分隔
            print("SAP_Nums_Search")
            SAP_Nums_List = re.split(r'[\s,;]+', SAP_Number_List.strip())
            print(SAP_Nums_List)

            sql_result = []
            for SAPNo_Searchby in SAP_Nums_List:
                sql_result_each, columnNameList = db.fetch(
                    tableName=tableName, 
                    dbindex=dbindex, 
                    PartNo_Searchby='',
                    SAPNo_Searchby=SAPNo_Searchby,
                    PartValue_Searchby='',
                    MfcPartNum_Searchby='',
                    Description_Searchby='',
                    TechDescription_Searchby='',
                    Editor_Searchby='',
                    Manufacturer_Searchby=''
                    )
                if sql_result_each: # 非空结果才添加
                    sql_result.append(sql_result_each[0])
                else: # 未找到结果, 添加空行占位,填写SAP number
                    sql_result.append(['']*2 + [SAPNo_Searchby] + [''] * (len(columnNameList) - 3))
            sql_result_len = len(sql_result)
            return render_template(
                'index.html',
                Part_Type_List=Part_Type_List,
                sql_result=sql_result,
                columnNameList=columnNameList,  
                sql_result_len=sql_result_len,
                SAP_Nums=SAP_Number_List,
                MaxLine=MaxLine,
                Version=__Version__
                )
        else:
            return render_template('index.html', Part_Type_List=Part_Type_List, Version=__Version__)
    else:
        return render_template('index.html', Part_Type_List=Part_Type_List, Version=__Version__)

# =========== 以下部分为2026/01 之后新的AVL处理代码 ==============
def download_excel(output_excel_file, AJAX=False, msg_avlHandle="", btn_enabled=True):
    """ 弹窗下载AVL Excel文件。
    Args:
        param output_excel_file (str): 输出Excel文件路径
    return: 
        output_excel_file (str): 输出Excel文件路径
    """
    # 这里可以添加下载文件的逻辑
    if AJAX:
        # 生成Excel后
        download_url = url_for('downloadExcelFile', filename=os.path.basename(output_excel_file))
        return jsonify({'status': 'completed', 
                        'msg': msg_avlHandle, 
                        'btn_enabled': btn_enabled, 
                        'download_url': download_url
                        })
    else:
        response_file = send_file(output_excel_file, as_attachment=True)
        return response_file

def get_ordering_info_from_db(db, dbindex, tableName, SAP_Number_List, Multi_BOM_Info_list):
    """ 根据SAP编号列表,查询数据库,返回sql_result和columnNameList
    Args:
        db(db_mgt.Database): 数据库实例
        dbindex(int): 数据库索引
        tableName(str): 表名
        SAP_Number_List(list): SAP编号列表
        Multi_BOM_Info_list(list): BOM信息列表,用于查不到信息时, 获取SAP Description以便填写; 为空也不会出错

    Returns:
        tuple: (sql_result, columnNameList): 查询结果和列名列表
    """
    sql_result = []
    columnNameList = None
    for SAPNo_Searchby in SAP_Number_List:
        sql_result_each, columnNameList = db.fetch(
            tableName=tableName, 
            dbindex=dbindex, 
            PartNo_Searchby='',
            SAPNo_Searchby=SAPNo_Searchby,
            PartValue_Searchby='',
            MfcPartNum_Searchby='',
            Description_Searchby='',
            TechDescription_Searchby='',
            Editor_Searchby='',
            Manufacturer_Searchby=''
        )
        if sql_result_each:
            sql_result.append(sql_result_each[0])
        else:
            # 未找到结果, 添加空行占位,填写SAP number和Description
            SAP_Description_Searchby = ""
            for bom_info in Multi_BOM_Info_list:
                if bom_info.split(',')[0] == SAPNo_Searchby:
                    SAP_Description_Searchby = bom_info.split(',')[1]
                    break
            Not_Found_SAP_info = ('','', SAPNo_Searchby, SAP_Description_Searchby)
            sql_result.append(Not_Found_SAP_info)
    return sql_result, columnNameList

# 通过URL跳转的方式下载Excel文件
@app.route('/downloadExcelFile/<filename>')
def downloadExcelFile(filename):
    filePathName = os.path.join(os.path.dirname(os.path.realpath(sys.argv[0])), 'ExportFiles', filename)
    return send_file(filePathName, as_attachment=True)

@app.route("/AVLhandle", methods=['GET','POST'])
def AVLHandle():
    db = get_db()
    # test flash('msg test'), 需要进行手动清除
    # flash('Welcome to AVL Handle Page!')
    # 网页运行信息
    msg_avlHandle = ""
    # 按键使能状态
    btn_enabled = True
    # 处理POST请求, 即点击按钮之后,判断按键类型并处理
    
    if request.method == 'POST':
        # 处理POST请求
        # 定义全局变量
        global first_AVL_Output_File  # 声明为全局变量, 第一次生成的AVL Excel文件路径
        global AVL_Compare_Output_File  # 声明为全局变量, 对比后的AVL Excel文件路径
        # 获取public部分设置
        user = request.form.get('user')
        pwd = request.form.get('password')
        DB_Select = request.form.get('DB_Select')
        AVL_include = request.form.get("AVL_include")
        # 获取First AVL Generation部分设置
        PCBA_Part_Number_List = request.form.get('PCBA_Part_Number_List')
        btn = request.form.get('btn')
        # 获取AVL Comparison部分设置
        AVL_Cmp_range = request.form.get("AVL_Cmp_range")
        excel_file = request.files.get('excel_file')

        # debug print
        debug_print("user:", user)
        debug_print("pwd:", pwd)
        debug_print("DB_Select:", DB_Select)
        debug_print("AVL_include:", AVL_include)
        debug_print("PCBA_Part_Number_List:", PCBA_Part_Number_List)
        debug_print("btn:", btn)
        debug_print("AVL_Cmp_range:", AVL_Cmp_range)
        debug_print("excel_file:", excel_file)

        # 判断是否输入账号密码
        if (not user or user.strip() == "") or (not pwd or pwd.strip() == ""):
            msg_avlHandle = "Windchill user name and password cannot be empty. Please input valid credentials."
            return jsonify({
                'status': 'error', 
                'msg': msg_avlHandle,
                'btn_enabled': btn_enabled
            })


        # 以下代码用于处理按钮点击后,界面显示和按钮使能状态,已经在页面的JavaScript中实现,这里注释掉
        # disable all buttons during processing
        # btn_enabled = False  
        # msg_avlHandle = "正在处理,请稍候..."
        
        # 处理Create AVL按钮点击事件
        if btn == 'Create_AVL':
            debug_print("="*30)
            debug_print("Create_AVL button clicked. Start processing...")
            # step1: 准备工作,处理输入参数
            # 判断PCBA Part Number List是否为空
            if not PCBA_Part_Number_List or PCBA_Part_Number_List.strip() == "":
                msg_avlHandle = "PCBA Part Number list cannot be empty. Please input valid PCBA Part Numbers."
                return jsonify({
                    'status': 'error', 
                    'msg': msg_avlHandle,
                    'btn_enabled': btn_enabled
                })
            # 处理 PCBA_Part_Number_List,获取list格式
            PCBA_Part_Number_List = re.split(r'[\s,;]+', PCBA_Part_Number_List.strip())
            # 处理AVL_include选项
            bCHINA_PN_ONLY = True if AVL_include == '2TFU CN only' else False
            # 变量定义
            Multi_BOM_Info_list = []    # PartNumber,PartName,Quantity,DesignatorRange
            Multi_BOM_SAP_Number_List = []  # SAP Number list in the BOM
            Multi_BOM_SAP_Number_List_Str = ""  # SAP Number list in the BOM, str format
            Multi_PCBA_Part_info_list = []  # PCBA Part info list
            # Step2: 连接PLM,获取BOM信息
            debug_print("Starting to get BOM from PLM...")
            for BOM_number in PCBA_Part_Number_List:
                print("Processing PCBA Part Number:", BOM_number)
                BOM_Info_list, BOM_SAP_Number_List, BOM_SAP_Number_List_Str, PCBA_Part_info_list, PLM_Login_OK = plm.get_BOM(user, pwd, BOM_number, bCHINA_PN_ONLY)
                if not PLM_Login_OK:
                    debug_print(f"Failed to login to Windchill for PCBA Part Number: {BOM_number}")
                    msg_avlHandle = "Failed to login to Windchill. Please check your username, password and network connection."
                    return jsonify({
                        'status': 'error', 
                        'msg': msg_avlHandle,
                        'btn_enabled': btn_enabled
                    })
                else:
                    debug_print(f"BOM Info for {BOM_number} retrieved successfully.")
                    Multi_BOM_Info_list += BOM_Info_list
                    Multi_BOM_SAP_Number_List += BOM_SAP_Number_List
                    Multi_BOM_SAP_Number_List_Str += BOM_SAP_Number_List_Str
                    Multi_PCBA_Part_info_list.append(PCBA_Part_info_list)
            # 去除重复项
            Multi_BOM_SAP_Number_List = list(set(Multi_BOM_SAP_Number_List))
            Multi_BOM_Info_list = list(set(Multi_BOM_Info_list))
            # Step3: 通过SQL获取ordering information
            debug_print("Starting to get ordering info from DB...")
            tableName = '---All----' #检索所有表
            dbindex = int(DB_Select)    #0: CONNECT DB, 1: Access DB
            # 打开DB
            bIsDBOpen = db.openDB(dbindex, db_mgt.DBList, app)
            if bIsDBOpen == True:
                flash(db_mgt.DBList[dbindex]+" Database opened successfully!")
            else:
                flash(db_mgt.DBList[dbindex]+" Database open error")
                return jsonify({
                    'status': 'error',
                    'msg': "Failed to open the selected database.",
                    'btn_enabled': btn_enabled
                })
            # 调用重构后的函数，获取ordering information
            sql_result, columnNameList = get_ordering_info_from_db(
                db, dbindex, tableName, Multi_BOM_SAP_Number_List, Multi_BOM_Info_list
            )
            # 判断是否有查询到数据，若没有则不继续处理，并提示用户
            sql_result_len = len(sql_result)
            if sql_result_len == 0:
                msg_avlHandle = "No ordering information found for the given PCBA Part Numbers' BOM SAP Numbers."
                return jsonify({
                    'status': 'error', 
                    'msg': msg_avlHandle,
                    'btn_enabled': btn_enabled
                })
            # Step4: 保存Excel文件, 使用模板2TFP900033A1076.xlsx
            debug_print("Starting to write AVL Excel file...")
            # Excel模板路径
            template_file = os.path.join(os.path.dirname(os.path.realpath(sys.argv[0])),  '2TFP900033A1076.xlsx')
            # 输出文件路径
            first_AVL_Output_File = os.path.join(os.path.dirname(os.path.realpath(sys.argv[0])), 'ExportFiles', f"AVL_Result_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
            # 调用Excel处理模块,生成AVL Excel文件, 结果直接输出到output_file
            excel_handle.first_write_AVL_to_excel(template_file, sql_result, Multi_PCBA_Part_info_list, first_AVL_Output_File)
            # Step5: 完成操作,返回结果信息
            debug_print("AVL Excel file created successfully.")
            # Step6: 提供下载
            msg_avlHandle = "Create AVL button processing completed. If the save dialog did not pop up, please click the Download_AVL button."
            # AJAX方式下载文件
            return download_excel(first_AVL_Output_File, AJAX=True, msg_avlHandle=msg_avlHandle, btn_enabled=True)
        # 处理Download AVL按钮点击事件
        elif btn == 'Download_AVL':
            debug_print("="*30)
            debug_print("Download_AVL button clicked. Start processing...")
            if not first_AVL_Output_File or not os.path.exists(first_AVL_Output_File):
                msg_avlHandle = "AVL Excel file not found. Please click the Create AVL button first."
                return jsonify({
                    'status': 'error', 
                    'msg': msg_avlHandle,
                    'btn_enabled': btn_enabled
                })
            # AJAX方式下载文件
            debug_print("Download AVL successfully.")
            msg_avlHandle = "Download_AVL button processing completed."
            return download_excel(first_AVL_Output_File, AJAX=True, msg_avlHandle=msg_avlHandle, btn_enabled=True)
        # 处理Manual Compare AVL按钮点击事件
        elif btn == 'Compare_Manual_AVL':
            debug_print("="*30)
            debug_print("Compare_Manual_AVL button clicked. Start processing...")
            # step1: 准备工作,处理输入参数
            # 判断excel_file是否为空
            if not excel_file:
                msg_avlHandle = "Please upload an Excel file for AVL comparison."
                return jsonify({
                    'status': 'error', 
                    'msg': msg_avlHandle,
                    'btn_enabled': btn_enabled
                })
            # 保存上传的Excel文件到临时路径
            temp_dir = tempfile.gettempdir()
            filename_with_ext = secure_filename(excel_file.filename)
            filename_no_ext = os.path.splitext(filename_with_ext)[0]
            uploaded_file = os.path.join(temp_dir, filename_with_ext)
            excel_file.save(uploaded_file)
            debug_print("Uploaded Excel file saved to:", uploaded_file)
            # 判断Excel文件中是否包含"AVL"和"AVL_Cmp"两个sheet
            required_sheets = excel_handle.AVL_MANUAL_REQUIRED_SHEETS
            if not excel_handle.check_AVL_file(uploaded_file, required_sheets):
                msg_avlHandle = f"The uploaded Excel file must contain the following sheets: {', '.join(required_sheets)}."
                return jsonify({
                    'status': 'error', 
                    'msg': msg_avlHandle,
                    'btn_enabled': btn_enabled
                })
            # 输出文件路径 
            AVL_Compare_Output_File = os.path.join(os.path.dirname(os.path.realpath(sys.argv[0])), 'ExportFiles', f"{filename_no_ext}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
            excel_handle.compare_avl_sheets(uploaded_file, AVL_Compare_Output_File)
            msg_avlHandle = "Compare Manual AVL button processing completed. If the save dialog did not pop up, please click the Download_Result button."
            # AJAX方式下载文件
            return download_excel(AVL_Compare_Output_File, AJAX=True, msg_avlHandle=msg_avlHandle, btn_enabled=True)
        # 处理Compare AVL按钮点击事件
        elif btn == 'Compare_AVL':
            debug_print("="*30)
            debug_print("Compare_AVL button clicked. Start processing...")
            # step1: 准备工作,处理输入参数
            # 判断excel_file是否为空
            if not excel_file:
                msg_avlHandle = "Please upload an Excel file for AVL comparison."
                return jsonify({
                    'status': 'error', 
                    'msg': msg_avlHandle,
                    'btn_enabled': btn_enabled
                })
            # 保存上传的Excel文件到临时路径
            temp_dir = tempfile.gettempdir()
            filename_with_ext = secure_filename(excel_file.filename)
            filename_no_ext = os.path.splitext(filename_with_ext)[0]
            uploaded_file = os.path.join(temp_dir, filename_with_ext)
            excel_file.save(uploaded_file)
            debug_print("Uploaded Excel file saved to:", uploaded_file)
            # 判断Excel文件中是否包含"AVL" sheet
            required_sheets = excel_handle.AVL_AUTO_REQUIRED_SHEETS
            if not excel_handle.check_AVL_file(uploaded_file, required_sheets):
                msg_avlHandle = f"The uploaded Excel file must contain the following sheet: {', '.join(required_sheets)}."
                return jsonify({
                    'status': 'error', 
                    'msg': msg_avlHandle,
                    'btn_enabled': btn_enabled
                })
            # step2: 根据Compare范围, 连接数据库获取sql_result
            debug_print("Starting to create AVL_Cmp sheet...")
            # Excel模板路径
            template_file = os.path.join(os.path.dirname(os.path.realpath(sys.argv[0])),  '2TFP900033A1076.xlsx')
            # todo: 1. 根据AVL_Cmp_range选择不同的处理方式, 生成AVL_Cmp sheet
            # 选项1: AVL Sheet Only, 获取上传文件中的AVL sheet内容，然后根据B列第7行开始的Part Numbers查询数据库，得到结果待用
            if AVL_Cmp_range == 'AVL_Sheet_Only':
                # Step A1: 获取SAP Numbers列表
                SAP_Nums_List = excel_handle.get_SAP_Numbers_from_AVL_sheet(uploaded_file)
                debug_print("SAP Numbers extracted from AVL sheet:", SAP_Nums_List)
                debug_print("len(SAP_Nums_List):", len(SAP_Nums_List))
                # 判断是否有SAP Numbers
                if len(SAP_Nums_List) == 0:
                    msg_avlHandle = "No SAP Numbers found in the uploaded AVL sheet."
                    return jsonify({
                        'status': 'error', 
                        'msg': msg_avlHandle,
                        'btn_enabled': btn_enabled
                    })
                # Step A2: 获取ordering information
                debug_print("Starting to get ordering info from DB...")
                tableName = '---All----' #检索所有表
                dbindex = int(DB_Select)    #0: CONNECT DB, 1: Access DB
                # 打开DB
                bIsDBOpen = db.openDB(dbindex, db_mgt.DBList, app)
                if bIsDBOpen == True:
                    flash(db_mgt.DBList[dbindex]+" Database opened successfully!")
                else:
                    flash(db_mgt.DBList[dbindex]+" Database open error")
                    return jsonify({
                        'status': 'error',
                        'msg': "Failed to open the selected database.",
                        'btn_enabled': btn_enabled
                    })
                # 调用重构后的函数，获取ordering information
                sql_result, columnNameList = get_ordering_info_from_db(db, dbindex, tableName, SAP_Nums_List, []
                )   # Multi_BOM_Info_list为空也不会出错
            # 选项2: BOM Related sheet,  获取上传文件中的BOM Related sheet内容，然后根据B列第3行开始的PCBA Part Numbers查询PLM获取BOM，然后根据BOM中的SAP Numbers查询数据库，生成AVL_Cmp sheet
            elif AVL_Cmp_range == 'BOM_Related_Sheet':
                # 此部分的功能与Create AVL部分类似, 但是没有重构
                # Step B1: 从upload file中获取PCBA Part Numbers列表
                PCBA_Part_Number_List = excel_handle.get_PCBA_Part_Numbers_from_BOM_Related_sheet(uploaded_file)
                # 以下部分直接从Create AVL中复制过来, 需要重构
                # 处理AVL_include选项
                bCHINA_PN_ONLY = True if AVL_include == '2TFU CN only' else False
                # 变量定义
                Multi_BOM_Info_list = []    # PartNumber,PartName,Quantity,DesignatorRange
                Multi_BOM_SAP_Number_List = []  # SAP Number list in the BOM
                Multi_BOM_SAP_Number_List_Str = ""  # SAP Number list in the BOM, str format
                Multi_PCBA_Part_info_list = []  # PCBA Part info list
                # Step2: 连接PLM,获取BOM信息
                debug_print("Starting to get BOM from PLM...")
                for BOM_number in PCBA_Part_Number_List:
                    print("Processing PCBA Part Number:", BOM_number)
                    BOM_Info_list, BOM_SAP_Number_List, BOM_SAP_Number_List_Str, PCBA_Part_info_list, PLM_Login_OK = plm.get_BOM(user, pwd, BOM_number, bCHINA_PN_ONLY)
                    if not PLM_Login_OK:
                        debug_print(f"Failed to login to Windchill for PCBA Part Number: {BOM_number}")
                        msg_avlHandle = "Failed to login to Windchill. Please check your username, password and network connection."
                        return jsonify({
                            'status': 'error', 
                            'msg': msg_avlHandle,
                            'btn_enabled': btn_enabled
                        })
                    else:
                        debug_print(f"BOM Info for {BOM_number} retrieved successfully.")
                        Multi_BOM_Info_list += BOM_Info_list
                        Multi_BOM_SAP_Number_List += BOM_SAP_Number_List
                        Multi_BOM_SAP_Number_List_Str += BOM_SAP_Number_List_Str
                        Multi_PCBA_Part_info_list.append(PCBA_Part_info_list)
                # 去除重复项
                Multi_BOM_SAP_Number_List = list(set(Multi_BOM_SAP_Number_List))
                Multi_BOM_Info_list = list(set(Multi_BOM_Info_list))
                # Step3: 通过SQL获取ordering information
                debug_print("Starting to get ordering info from DB...")
                tableName = '---All----' #检索所有表
                dbindex = int(DB_Select)    #0: CONNECT DB, 1: Access DB
                # 打开DB
                bIsDBOpen = db.openDB(dbindex, db_mgt.DBList, app)
                if bIsDBOpen == True:
                    flash(db_mgt.DBList[dbindex]+" Database opened successfully!")
                else:
                    flash(db_mgt.DBList[dbindex]+" Database open error")
                    return jsonify({
                        'status': 'error',
                        'msg': "Failed to open the selected database.",
                        'btn_enabled': btn_enabled
                    })
                # 调用重构后的函数，获取ordering information
                sql_result, columnNameList = get_ordering_info_from_db(
                    db, dbindex, tableName, Multi_BOM_SAP_Number_List, Multi_BOM_Info_list
                )
                # 判断是否有查询到数据，若没有则不继续处理，并提示用户
                sql_result_len = len(sql_result)
                if sql_result_len == 0:
                    msg_avlHandle = "No ordering information found for the given PCBA Part Numbers' BOM SAP Numbers."
                    return jsonify({
                        'status': 'error', 
                        'msg': msg_avlHandle,
                        'btn_enabled': btn_enabled
                    })
            # 以下部分为通用部分
            # Step 3： 输出AVL Comparison Excel文件
            # 判断是否有查询到数据，若没有则不继续处理，并提示用户
            # 不需要，因为SAP Numbers已经判断过了
            debug_print("Starting to write AVL Comparison Excel file...")
            # 输出文件路径
            AVL_Compare_Output_File = os.path.join(os.path.dirname(os.path.realpath(sys.argv[0])), 'ExportFiles', f"{filename_no_ext}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
            # 调用Excel处理模块，生成AVL Comparison Excel文件, 结果直接输出到output_file
            # 将sql_result写入到临时文件
            tmp_file = os.path.join(temp_dir, f"temp_sql_result_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
            excel_handle.first_write_AVL_to_excel(template_file, sql_result, [], tmp_file)
            # 生成包含AVL,AVL_Cmp两个sheet的Excel文件
            excel_handle.copy_AVL_to_AVL_Cmp_In_UploadFile(tmp_file, uploaded_file)
            # 在uploaded_file的基础上，生成AVL_Cmp结果文件
            excel_handle.compare_avl_sheets(uploaded_file, AVL_Compare_Output_File)

            # Step5: 完成操作,返回结果信息
            debug_print("Compare AVL successfully.")
            msg_avlHandle = "Create AVL button processing completed. If the save dialog did not pop up, please click the Download_Result button."
            return download_excel(AVL_Compare_Output_File, AJAX=True, msg_avlHandle=msg_avlHandle, btn_enabled=True)
        # 处理Download Result按钮点击事件
        elif btn == 'Download_Result':
            debug_print("="*30)
            debug_print("Download_Result button clicked. Start processing...")
            if not AVL_Compare_Output_File or not os.path.exists(AVL_Compare_Output_File):
                msg_avlHandle = "AVL Comparison Excel file not found. Please click the Compare AVL button first."
                return jsonify({
                    'status': 'error', 
                    'msg': msg_avlHandle,
                    'btn_enabled': btn_enabled
                })
            debug_print("Download Result successfully.")
            msg_avlHandle = "Download_Result button processing completed."
            return download_excel(AVL_Compare_Output_File, AJAX=True, msg_avlHandle=msg_avlHandle, btn_enabled=True)
        # 处理未知按钮点击事件
        else:
            debug_print("="*30)
            debug_print("Unknown button clicked. Start processing...")
            flash("未知的操作按钮！")
            # 若出现此情况,直接返回页面
            return render_template('AVLHandle.html',
                                   user=user,
                                   pwd=pwd,
                                   PCBA_Part_Number_List=PCBA_Part_Number_List,
                                   Version=__Version__,
                                   msg_avlHandle=msg_avlHandle,
                                   btn_enabled=btn_enabled)
        # Flask常规返回方法,因使用AJAX,此处注释掉
        # return render_template('AVLHandle.html',
        #                        user=user,
        #                        pwd=pwd,
        #                        PCBA_Part_Number_List=PCBA_Part_Number_List,
        #                        Version=__Version__,
        #                        msg_avlHandle=msg_avlHandle,
        #                        btn_enabled=btn_enabled)
        # enable buttons after processing
        btn_enabled = True
        # 对按键响应操作完成,返回JSON以便AJAX处理
        # JavaScript方式刷新页面
        return jsonify({'status': 'completed',
                         'msg': msg_avlHandle,
                         'btn_enabled': btn_enabled})
    return render_template('AVLHandle.html',
                           Version=__Version__,
                           msg_avlHandle=msg_avlHandle,
                           btn_enabled=btn_enabled)

# =========== 以下部分Database同步情况自动检查代码 ==============
def get_DBSync_Sql_Result_By_SAP_Number(DB_Select, SAP_Numbers_List):
    """ 根据SAP Numbers列表,查询数据库,返回sql_CONNECT_result, sql_DB_result, columnNameList, msg
    Args:
        DB_Select(str): 数据库选择
        SAP_Numbers_List(list): SAP Numbers列表
    Returns:
        sql_CONNECT_result, sql_DB_result, columnNameList, msg
    """
    sql_CONNECT_result = []
    sql_DB_result = []
    columnNameList = None
    msg = ""
    tableName = '---All----' #检索所有表
    # 线程安全：每次新建db实例
    db_local = db_mgt.Database()
    # 先查询CONNECT DB作为参考数据
    dbindex = 0    #0: CONNECT DB, 1: Access DB
    bIsDBOpen = db_local.openDB(dbindex, db_mgt.DBList, app)
    if bIsDBOpen == True:
        debug_print(db_mgt.DBList[dbindex]+" Database opened successfully!")
    else:
        debug_print(db_mgt.DBList[dbindex]+" Database open error")
        msg = "Failed to open the CONNECT database."
        return sql_CONNECT_result, sql_DB_result, columnNameList, msg
    for SAPNo_Searchby in SAP_Numbers_List:
        sql_result_each, columnNameList = db_local.fetch(
            tableName=tableName, 
            dbindex=dbindex, 
            PartNo_Searchby='',
            SAPNo_Searchby=SAPNo_Searchby,
            PartValue_Searchby='',
            MfcPartNum_Searchby='',
            Description_Searchby='',
            TechDescription_Searchby='',
            Editor_Searchby='',
            Manufacturer_Searchby=''
        )
        if sql_result_each:
            sql_CONNECT_result.append(sql_result_each[0])
        else:
            Not_Found_Part_info = ('','', SAPNo_Searchby)
            sql_CONNECT_result.append(Not_Found_Part_info)
    # 再查询Access DB
    dbindex = int(DB_Select)    # 2: CNILG Access DB, 3: CNILX Access DB
    bIsDBOpen = db_local.openDB(dbindex, db_mgt.DBList, app)
    if bIsDBOpen == True:
        debug_print(db_mgt.DBList[dbindex]+" Database opened successfully!")
    else:
        debug_print(db_mgt.DBList[dbindex]+" Database open error")
        msg = "Failed to open the selected Access database."
        return sql_CONNECT_result, sql_DB_result, columnNameList, msg
    for SAPNo_Searchby in SAP_Numbers_List:
        sql_result_each, columnNameList = db_local.fetch(
            tableName=tableName, 
            dbindex=dbindex, 
            PartNo_Searchby='',
            SAPNo_Searchby=SAPNo_Searchby,
            PartValue_Searchby='',
            MfcPartNum_Searchby='',
            Description_Searchby='',
            TechDescription_Searchby='',
            Editor_Searchby='',
            Manufacturer_Searchby=''
        )
        if sql_result_each:
            sql_DB_result.append(sql_result_each[0])
        else:
            Not_Found_Part_info = ('','', SAPNo_Searchby)
            sql_DB_result.append(Not_Found_Part_info)
    msg = "Get database result successfully."
    return sql_CONNECT_result, sql_DB_result, columnNameList, msg

def cmp_DBSync_Result(sql_CONNECT_result, sql_DB_result, columnNameList, compare_columns=None):
    """ 比较两个数据库的查询结果, 结果输出到Excel文件,间隔行显示差异,对比完成返回Excel文件路径
    Args:
        sql_CONNECT_result(list): CONNECT数据库查询结果
        sql_DB_result(list): 目标数据库查询结果
        columnNameList(list): 列名列表
    Returns:
        dbsyncinfo(str): 数据库同步比较结果字符串
        cmpare_excel_file(str): 输出Excel文件路径
        diff_count(int): 差异数量

    """
    # 输出Excel文件路径
    output_excel_file = os.path.join(os.path.dirname(os.path.realpath(sys.argv[0])), 'ExportFiles', f"DB_Sync_Comparison_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DB Sync Comparison"
    # 写入列名（右移一列）
    ws.cell(row=1, column=1, value="Item")
    for col_index, col_name in enumerate(columnNameList, start=2):
        ws.cell(row=1, column=col_index, value=col_name)
    # 比较结果写入Excel
    max_rows = max(len(sql_CONNECT_result), len(sql_DB_result))
    diff_count = 0
    # 计算需要对比的列索引
    if compare_columns is not None:
        compare_indices = [columnNameList.index(col) for col in compare_columns if col in columnNameList]
    else:
        compare_indices = list(range(len(columnNameList)))

    for row_index in range(max_rows):
        connect_row = sql_CONNECT_result[row_index] if row_index < len(sql_CONNECT_result) else [''] * len(columnNameList)
        db_row = sql_DB_result[row_index] if row_index < len(sql_DB_result) else [''] * len(columnNameList)
        seq_num = row_index + 1
        # 写入CONNECT数据库行（右移一列，A列序号）
        ws.cell(row=row_index * 2 + 2, column=1, value=seq_num)
        for col_index, cell_value in enumerate(connect_row, start=2):
            ws.cell(row=row_index * 2 + 2, column=col_index, value=cell_value)
        # 写入目标数据库行（右移一列，A列序号）
        ws.cell(row=row_index * 2 + 3, column=1, value=seq_num)
        for col_index, cell_value in enumerate(db_row, start=2):
            ws.cell(row=row_index * 2 + 3, column=col_index, value=cell_value)
        # 单元格逐个对比（右移一列）
        row_diff = False
        # 对比并设置底色
        green_fill = openpyxl.styles.PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")  # 浅绿
        red_fill = openpyxl.styles.PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")  # 红色
        if compare_columns is None:
            # 保持原有方式：所有列都对比
            for col_index in compare_indices:
                val1 = connect_row[col_index] if col_index < len(connect_row) else ''
                val2 = db_row[col_index] if col_index < len(db_row) else ''
                if val1 != val2:
                    row_diff = True
                    ws.cell(row=row_index * 2 + 2, column=col_index + 2).fill = red_fill
                    ws.cell(row=row_index * 2 + 3, column=col_index + 2).fill = red_fill
                else:
                    ws.cell(row=row_index * 2 + 2, column=col_index + 2).fill = green_fill
                    ws.cell(row=row_index * 2 + 3, column=col_index + 2).fill = green_fill
        else:
            # 只对比指定列
            for col_index in compare_indices:
                val1 = connect_row[col_index] if col_index < len(connect_row) else ''
                val2 = db_row[col_index] if col_index < len(db_row) else ''
                if val1 != val2:
                    row_diff = True
                    ws.cell(row=row_index * 2 + 2, column=col_index + 2).fill = red_fill
                    ws.cell(row=row_index * 2 + 3, column=col_index + 2).fill = red_fill
                else:
                    ws.cell(row=row_index * 2 + 2, column=col_index + 2).fill = green_fill
                    ws.cell(row=row_index * 2 + 3, column=col_index + 2).fill = green_fill
        if row_diff:
            diff_count += 1
    wb.save(output_excel_file)
    dbsyncinfo = f"Database synchronization comparison completed. Total differences found: {diff_count}."
    return dbsyncinfo, output_excel_file, diff_count

def send_outlook_mail(to, subject, body, attachment_path=None):
    try:
        outlook = win32.Dispatch('outlook.application')
        mail = outlook.CreateItem(0)
        mail.To = to
        mail.Subject = subject
        mail.Body = body
        if attachment_path:
            mail.Attachments.Add(attachment_path)
        mail.Send()
    except Exception as e:
        debug_print(f"Failed to send mail: {e}")

def background_check(DB_Select, SAP_Numbers_List, Reminder_Email, Check_Interval_Time, MAX_TRY, Chk_Scope):
    try_count = 0
    # get chk scope
    if Chk_Scope == 'All_Columns':
        compare_columns = None
    elif Chk_Scope == 'eCAD_Related_Columns':
        # 仅比较eCAD相关列,exclude:parttype,editor,us_technology
        compare_columns = ['partnumber', 'value', 'sap_number', 'sap_description', 'parttype', 'manufact 1', 'manufact partnum 1', 'datasheet 1', 'manufact 2', 'manufact partnum 2', 'datasheet 2', 'manufact 3', 'manufact partnum 3', 'datasheet 3', 'manufact 4', 'manufact partnum 4', 'datasheet 4', 'manufact 5', 'manufact partnum 5', 'datasheet 5', 'manufact 6', 'manufact partnum 6', 'datasheet 6', 'manufact 7', 'manufact partnum 7', 'datasheet 7', 'scm_symbol', 'pcb_footprint', 'alt_symbols', 'mounttechn', 'ad_symbol', 'ad_footprint', 'ad_alt_footprint', 'detaildrawing', 'status',  'techdescription']
    while True:
        try_count += 1
        sql_CONNECT_result, sql_DB_result, columnNameList, msg = get_DBSync_Sql_Result_By_SAP_Number(DB_Select, SAP_Numbers_List)
        dbsyncinfo, cmpare_excel_file, diff_count = cmp_DBSync_Result(sql_CONNECT_result, sql_DB_result, columnNameList, compare_columns)
        debug_print(f"[DBSyncCheck] Try {try_count}, diff_count={diff_count}")
        if diff_count == 0:
            subject = f"DB Sync Check finished with No Differences"
            body = (
                f"{dbsyncinfo}\n\n当前差异计数(diff_count): 0\n数据库同步完成，自动对比结束。"
                f"\n\nChecked SAP Numbers: {', '.join(SAP_Numbers_List)}\nTime: {datetime.datetime.now()}"
            )
            send_outlook_mail(Reminder_Email, subject, body, cmpare_excel_file)
            debug_print("[DBSyncCheck] No differences found, stopping background check.")
            break
        elif try_count < MAX_TRY:
            subject = f"DB Sync Check Result (Try {try_count})"
            body = (
                f"{dbsyncinfo}\n\n当前差异计数(diff_count): {diff_count}"
                f"\n\nChecked SAP Numbers: {', '.join(SAP_Numbers_List)}\nTime: {datetime.datetime.now()}"
            )
            send_outlook_mail(Reminder_Email, subject, body, cmpare_excel_file)
        else:
            # 达到最大尝试次数，发送超时通知邮件
            subject = f"DB Sync Check Timeout (try_count >= {MAX_TRY})"
            body = f"DB同步检查超时，差异计数达到{diff_count}，已停止自动检查。\n\nChecked SAP Numbers: {', '.join(SAP_Numbers_List)}\nTime: {datetime.datetime.now()}"
            send_outlook_mail(Reminder_Email, subject, body, cmpare_excel_file)
            break
        time.sleep(Check_Interval_Time * 60)
    debug_print("[DBSyncCheck] Background check finished.")

@app.route("/dbsynccheck", methods=['GET','POST'])
def DBSyncCheck():
    #调用db_mgt中的函数,返回数据库同步情况
    # dbsyncinfo=db.dbSyncCheck(db_mgt.DBList,app)
    Check_Interval_Time_List = [1, 5, 10, 15, 30, 60]  # minutes
    MAX_TRY_List = [1, 2, 3, 5, 10, 15, 30]  # 可选最大尝试次数
    btn_enabled = True

    if request.method == 'POST':
        # 处理POST请求
        # 获取表单数据
        DB_Select = request.form.get('DB_Select')
        SAP_Numbers_List = request.form.get('SAP_Numbers_List')
        Reminder_Email = request.form.get('Reminder_Email')
        Check_Interval_Time = int(request.form.get('Check_Interval_Time'))
        MAX_TRY = int(request.form.get('MAX_TRY', 2))
        Chk_Scope = request.form.get('Chk_Scope')
        # debug print
        debug_print("DB_Select:", DB_Select)
        debug_print("SAP_Numbers_List:", SAP_Numbers_List)
        debug_print("Reminder_Email:", Reminder_Email)
        debug_print("Check_Interval_Time:", Check_Interval_Time)
        debug_print("MAX_TRY:", MAX_TRY)
        debug_print("Chk_Scope:", Chk_Scope)
        # 判断是否输入必要参数
        if (not SAP_Numbers_List or SAP_Numbers_List.strip() == ""):
            flash("SAP Numbers list cannot be empty. Please input valid SAP Numbers.")
        elif (not Reminder_Email or Reminder_Email.strip() == ""):
            flash("Reminder email cannot be empty. Please input a valid email address.")
        else:
            btn_enabled = False
            SAP_Numbers_List_Split = re.split(r'[\s,;]+', SAP_Numbers_List.strip())
            t = threading.Thread(target=background_check, args=(DB_Select, SAP_Numbers_List_Split, Reminder_Email, Check_Interval_Time, MAX_TRY, Chk_Scope), daemon=True)
            t.start()
            flash(f"后台定时检查已启动，每{Check_Interval_Time}分钟检查一次，最多尝试{MAX_TRY}次，结果将通过邮件发送到{Reminder_Email}。")
            # 设置按键不可用，防止多次提交
            btn_enabled = False
        return render_template('DBSyncCheck.html',
                                Check_Interval_Time_List = Check_Interval_Time_List,
                                MAX_TRY_List = MAX_TRY_List,
                                btn_enabled=btn_enabled,
                                SAP_Numbers_List=SAP_Numbers_List,
                                Reminder_Email=Reminder_Email,
                                Version=__Version__,
                                MAX_TRY=MAX_TRY,
                                Chk_Scope=Chk_Scope)

    return render_template('DBSyncCheck.html',
                           Check_Interval_Time_List = Check_Interval_Time_List,
                           MAX_TRY_List = MAX_TRY_List,
                           btn_enabled=btn_enabled,
                           Version=__Version__)


# =========== 以下部分为Cyrus 生成的AVL BOM相关代码 ==============
#函数,功能为读取Windhill的BOM表并去除重复。输入,Excel Sheet, WinChill返回的JSON,Level是指BOM结构上的层级,1为首层
def showBOM(sheet,subpart,level):
    db = get_db()
    if level > 1:
        #判断PartNumber是否已经存在于当前AVL中
        if not subpart["PartNumber"] in AVLPart_ListView:
            #判断是否software Part或者Dcoment Part,只有不是时才往下走
            if subpart["Part"]["@odata.type"] != "#PTC.ProdMgmt.ABBDOCPART" and subpart["Part"]["@odata.type"] != "#PTC.ProdMgmt.ABBSOFTWAREPART":
                #写入到partlist里
                AVLPart_ListView.add(subpart["PartNumber"])
                #写入到Excel里
                global Excel_Row #声明Excel_Row为全局变量
                rownum1="B"+str(Excel_Row)
                rownum2="C"+str(Excel_Row)
                sheet[rownum1]=subpart["PartNumber"]    #B列写入Windchill中的PartNumber
                sheet[rownum2]=subpart["PartName"]      #C列写入Windchill中的Description
                
                #print(subpart["PartNumber"])
                #######读取CONNECT数据库并返回内容
                #根据PartNumber读取CONNECT里的内容
                print(Excel_Row-6)
                Excel_Row+=1
                ConnectFetch=db.fetchMax(subpart["PartNumber"])
                
                #声明全球变量
                global Component_ListView
                if isinstance(ConnectFetch,list):       #只有在数据库中存在该值,如果的返回的值是一个list
                    if len(ConnectFetch)>0:             #返回和数据中,超过一行,即有有效数据
                        ConnectData=ConnectFetch[0]
                        #遍历返回的CONNECT数据库

                        for index in range(1,16):
                            rownum3=chr(index+67)+str(Excel_Row-1) #第一个字母为D,从D列开始往后写
                            if ConnectData[index]!="":
                                sheet[rownum3]=ConnectData[index+2] #按照db_mgt里的数值,将结果加上与Manufactory对上的列
                        coid=Excel_Row-7
                        Component_ListView.add(str(coid)+","+subpart["PartNumber"]+","+subpart["PartName"]+","+ConnectData[1]+","+ConnectData[2])
                else:
                    Component_ListView.add(str(coid)+","+subpart["PartNumber"]+","+subpart["PartName"]+",,")                        

    #print("Components" in subpart)
    if "Components" in subpart:
        if len(subpart["Components"])>0:
            for subpart2 in subpart["Components"]:
                showBOM(sheet,subpart2,level+1)
                
                
#函数,用于检验返回的值是否Json语句,以判断是否正确地访问windchill
def is_json(myjson):  
    try:  
        json_object = json.loads(myjson)  
    except ValueError as e:  
        return False  
    return True  


#avlindex页面,生成AVL的入口页面
@app.route("/avlindex", methods=['GET','POST'])
def AVLIndex():
    return render_template('AVLIndex.html')

#avl export页面,生成AVL后的返回页面
@app.route('/exportavl',methods=['GET','POST'])  
def exportavl():
    db = get_db()
    #由于运行在服务器,每次访问时,均需要先重置Global变量以达到预期效果
    global AVLPart_ListView
    AVLPart_ListView.clear()
    global Excel_Row
    Excel_Row=7
    global Component_ListView
    Component_ListView.clear()
 
    username = request.form.get('user')
    password = request.form.get('password')
    # 将用户名和密码组合成一个字符串,并用冒号分隔  
    credentials = f"{username}:{password}"  
    # 对这个字符串进行base64编码  
    encoded_credentials = base64.b64encode(credentials.encode('utf-8'))  
    
    partnumber = request.form.get('partnumber')
    #print(partnumber)
    ########第一步,打开Excel文件并用于数据中转
    # 加载现有的 Excel 文件  
    workbook = openpyxl.load_workbook('2TFP900033A1076.xlsx')  
    #指定sheet为AVL
    sheet1 = workbook["BOM Related"]   


    ########第二步,获取WindChill Token
    url = WC_Path + 'PTC/GetCSRFToken()'  # 目标 URL  
    headers = {  
        'Authorization': 'Basic ' + encoded_credentials.decode('utf-8'),  
        'Accept': 'application/json'  
    }  
    response = requests.get(url, headers=headers)  # 发送带请求头的 GET 请求  
    #如果返回值不为JSON,重新填写
    if not is_json(response.text):
        return render_template('AVLIndex.html', ErrorMessage="访问WindChill失败,请检查用户名、密码及网络连接")
    json_data = json.loads(response.text)  
    nonce_value = json_data.get('NonceValue')  
    headers['CSRF_NONCE'] = nonce_value

    ########第三步,打开ACCESS数据库并读取AVL对应的BOM表
    #打开数据库
    bIsDBOpen=db.openAcc()
    #打开数据表,并查找对应的
    sql_result=db.readBOM(partnumber)
    
    #定义一个空的集合用于显示BOM清单
    BOM_ListView=set()
    
    #打开CONNECT数据库
    db.openMaxDB()
    
    #遍历所有sql_result
    for index in range(len(sql_result)):
        
        
        # 修改单元格内容
        rownum = 'B' + str(index+3)
        print(index)
        BomNum = re.sub(r'[\',()]', '', str(sql_result[index])) 
        sheet1[rownum] = BomNum
        
        ########第四步,从WindChill里导入BOM表的状态
        url = WC_Path + f"ProdMgmt/Parts?$filter=Number eq '{BomNum}'"  # 目标 URL  
        response = requests.get(url, headers=headers)  # 发送带请求头的 GET 请求  
        json_data = json.loads(response.text)  
        partID=""
        # 遍历value数组中的每个元素  
        for partvalue in json_data['value']:  
            # 检查'View'字段是否为'design'（不区分大小写）  
            if partvalue['View'].lower() == 'design':  
                # 创建一个字典来存储part的详细信息  
                BOM_ListView.add(BomNum+","+ partvalue['State']['Value']+","+ partvalue['Version']+","+  partvalue['Name'])
                partID = partvalue["ID"]
                rownum = 'C' + str(index+3)
                sheet1[rownum] = partvalue['Name']
                # 由于已经找到了匹配的'design'视图,因此跳出循环  
                break
        
        #第五步,根据第四步返回的ID,查找对应的BOM结构
        url= WC_Path + "ProdMgmt/Parts('" + partID + "')/PTC.ProdMgmt.GetBOM?$expand=Components($expand=Part($select=Name,Number);$levels=max)"
        response = requests.post(url, headers=headers)  # 发送带请求头的 GET 请求 
        json_data = json.loads(response.text)          
        showBOM(workbook["AVL"],json_data,1)
    
    
    
    # 保存修改后的工作簿  
    workbook.save(filename="out/modified_example.xlsx")      
    PartCount=Excel_Row-7
    
    return render_template('AVLoutput.html', sql_result=BOM_ListView,AVL=partnumber,PartCount=PartCount,componentlist=Component_ListView)

#超链接,用于下载相应的Excel文件
@app.route('/downloadexcel/<AVL>')  
def downloadFile(AVL):  
    # 返回修改后的Excel文件供下载  
    modified_file = open("out/modified_example.xlsx", "rb")  
    
    # 保存并准备下载  
    response = send_file(modified_file, download_name=AVL+'.xlsx', as_attachment=True)  
    
    return response

if __name__ == '__main__':  
    app.run(host="0.0.0.0", debug = True)





