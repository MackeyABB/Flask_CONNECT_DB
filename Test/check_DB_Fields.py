

'''
Script Name: check_DB_Fields.py

Description:
This script connects independently to three databases (SAPMaxDB, AccessDB, PostgreSQL via ODBC), retrieves all field names for each table listed in the corresponding table lists, and outputs the results to Excel files. Each sheet represents a table (using the SAPMaxDB table name as the sheet name), and each column contains the sorted field names from the three databases for that table. The script does not depend on any other Python files in the project.

Version:
xx.yy.zz
xx: Major version (architecture changes)
yy: Feature additions
zz: Bug fixes
'''

__version__ = "1.0.0"


from typing import List
import pypyodbc
import openpyxl

# Table lists to be queried
# SAPMaxDB table list (for 01-Cadence CIS DB, ODBC, DESTO)
TABLES_SAPMaxDB: List[str] = [
    "CAPACITORS",
    "RESISTORS",
    "VARISTORS",
    "TRANSISTORS",
    "DIODES",
    "ICS_DIGITAL",
    "MEMORY",
    "ICS_ANALOG",
    "REGULATORS",
    "CONVERTERS",
    "OP_AMPS",
    "MAGNETICS",
    "TRANSFORMERS",
    "OPTO",
    "OSCILLATORS",
    "CONNECTORS",
    "RELAYS",
    "SENSORS",
    "SWITCHES",
    "MECHPARTS",
    "MISCPARTS",
    "PCB",
    "SOFTWARE",
    "TITLEBLOCK",
    "SHAPES"
]
# AccessDB table list (for 02-Altium Access DB, ODBC, CNILG)
TABLES_AccessDB: List[str] = [
    "[01-Capacitors]",
    "[02-Resistors]",
    "[03-Varistors]",
    "[04-Transistors]",
    "[05-Diodes]",
    "[06-ICs_digital]",
    "[07-Memory]",
    "[08-ICs_analog]",
    "[09-Regulators]",
    "[10-Converters]",
    "[11-OP_Amps]",
    "[12-Magnetics]",
    "[13-Transformers]",
    "[14-Opto]",
    "[15-Oscillators]",
    "[16-Connectors]",
    "[17-Relays]",
    "[18-Sensors]",
    "[19-Switches]",
    "[20-MechParts]",
    "[21-MiscParts]",
    "[23-PCB]",
    "[24-Software]",
    "[90-Titleblock]",
    "[98-Shapes]"
]

# PostgreSQL DB table list (for 03-PostgreSQL DB, ODBC, DESTO)
TABLES_PostgreSQLDB: List[str] = [
    "capacitors_l2u",
    "resistors_l2u",
    "varistors_l2u",
    "transistors_l2u",
    "diodes_l2u",
    "ics_digital_l2u",
    "memory_l2u",
    "ics_analog_l2u",
    "regulators_l2u",
    "converters_l2u",
    "op_amps_l2u",
    "magnetics_l2u",
    "transformers_l2u",
    "opto_l2u",
    "oscillators_l2u",
    "connectors_l2u",
    "relays_l2u",
    "sensors_l2u",
    "switches_l2u",
    "mechparts_l2u",
    "miscparts_l2u",
    "pcb_l2u",
    "software_l2u",
    "titleblock_l2u",
    "shapes_l2u"
]



def get_fields_odbc(conn_str, table_list):
    fields_dict = {}
    try:
        conn = pypyodbc.connect(conn_str, timeout=20, readonly=True)
        cursor = conn.cursor()
        for table in table_list:
            try:
                cursor.execute(f"SELECT * FROM {table} WHERE 1=0")
                fields = [desc[0] for desc in cursor.description]
                fields_dict[table] = fields
            except Exception as e:
                fields_dict[table] = [f"Error: {e}"]
        cursor.close()
        conn.close()
    except Exception as e:
        for table in table_list:
            fields_dict[table] = [f"DB Connect Error: {e}"]
    return fields_dict



def write_to_excel(filename, sap_table_names, sap_fields, access_fields, pg_fields):
    wb = openpyxl.Workbook()
    # Remove the default sheet
    if 'Sheet' in wb.sheetnames:
        std = wb['Sheet']
        wb.remove(std)
    for idx, sheet in enumerate(sap_table_names):
        # Excel sheet names cannot contain special characters
        safe_sheet = sheet.replace('[','').replace(']','')[:31]  # 31 character limit
        ws = wb.create_sheet(title=safe_sheet)
        ws.append(["TABLES_SAPMaxDB", "TABLES_AccessDB", "TABLES_PostgreSQLDB"])
        # Output the corresponding table names in the second row
        ws.append([
            TABLES_SAPMaxDB[idx] if idx < len(TABLES_SAPMaxDB) else "",
            TABLES_AccessDB[idx] if idx < len(TABLES_AccessDB) else "",
            TABLES_PostgreSQLDB[idx] if idx < len(TABLES_PostgreSQLDB) else ""
        ])
        # Get and sort field names
        sap_list = sap_fields.get(TABLES_SAPMaxDB[idx], [])
        access_list = access_fields.get(TABLES_AccessDB[idx], [])
        pg_list = pg_fields.get(TABLES_PostgreSQLDB[idx], [])
        # If not an error message, sort the field names
        sap_sorted = sorted(sap_list) if sap_list and not (len(sap_list)==1 and sap_list[0].startswith('Error')) and not (len(sap_list)==1 and sap_list[0].startswith('DB Connect Error')) else sap_list
        access_sorted = sorted(access_list) if access_list and not (len(access_list)==1 and access_list[0].startswith('Error')) and not (len(access_list)==1 and access_list[0].startswith('DB Connect Error')) else access_list
        pg_sorted = sorted(pg_list) if pg_list and not (len(pg_list)==1 and pg_list[0].startswith('Error')) and not (len(pg_list)==1 and pg_list[0].startswith('DB Connect Error')) else pg_list
        max_len = max(len(sap_sorted), len(access_sorted), len(pg_sorted))
        for i in range(max_len):
            row = [
                sap_sorted[i] if i < len(sap_sorted) else "",
                access_sorted[i] if i < len(access_sorted) else "",
                pg_sorted[i] if i < len(pg_sorted) else ""
            ]
            ws.append(row)
    wb.save(filename)

if __name__ == "__main__":
    # 1. Connection strings and parameters
    conn_str_sap = "DSN=CIS_Local;Uid=LIMBAS2USER;Pwd=LIMBASREAD;"
    conn_str_access = "DSN=CIS_PartLib_P_64;Uid=cadence_port;Pwd=Cadence_CIS.3;"
    conn_str_pg = "DSN=Connect ePDMS ODBC;Uid=odbc_user;Pwd=CONNECT2READ;"  # PostgreSQL ODBC connection info

    # 2. Retrieve fields
    print("Retrieving SAPMaxDB fields...")
    sap_fields = get_fields_odbc(conn_str_sap, TABLES_SAPMaxDB)
    print("Retrieving AccessDB fields...")
    access_fields = get_fields_odbc(conn_str_access, TABLES_AccessDB)
    print("Retrieving PostgreSQL fields...")
    pg_fields = get_fields_odbc(conn_str_pg, TABLES_PostgreSQLDB)

    # 3. Write to Excel
    print("Writing to Excel...")
    write_to_excel("Test/DB_Fields_Compare_SAPMaxDB.xlsx", TABLES_SAPMaxDB, sap_fields, access_fields, pg_fields)
    write_to_excel("Test/DB_Fields_Compare_AccessDB.xlsx", TABLES_AccessDB, sap_fields, access_fields, pg_fields)
    write_to_excel("Test/DB_Fields_Compare_PostgreSQLDB.xlsx", TABLES_PostgreSQLDB, sap_fields, access_fields, pg_fields)
    print("Done! Output files: Test/DB_Fields_Compare_SAPMaxDB.xlsx, Test/DB_Fields_Compare_AccessDB.xlsx, Test/DB_Fields_Compare_PostgreSQLDB.xlsx")
    


